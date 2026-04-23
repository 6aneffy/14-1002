from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from pypdf import PdfReader, PdfWriter

from .schema import Receipt
from .thumbnail import make_full_image_png_bytes

COLUMNS = [
    ("No.", 6),
    ("운임일자", 14),
    ("교통편", 16),
    ("등급", 14),
    ("매수", 8),
    ("금액(원)", 14),
    ("출발지", 14),
    ("도착지", 14),
    ("원본", 60),
]

IMAGE_COL_IDX = len(COLUMNS)  # "원본" 컬럼 인덱스 (1-based)
PX_TO_POINTS = 0.75  # 1 pt = 1/72 inch, 기본 96 DPI 가정
IMAGE_SCALE = 1 / 3  # 원본 대비 축소 비율


def build_xlsx(
    receipts: list[Receipt],
    files_map: dict[str, bytes] | None = None,
) -> bytes:
    """정산신청 리스트 xlsx 생성. files_map으로 원본 이미지를 셀에 삽입."""
    wb = Workbook()
    ws = wb.active
    ws.title = "정산신청"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    for idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = width

    files_map = files_map or {}

    for i, r in enumerate(receipts, start=1):
        row_idx = i + 1
        row = [
            i,
            r.travel_date.isoformat() if r.travel_date else "",
            r.transport.value,
            r.travel_class or "",
            r.quantity,
            r.amount,
            r.origin or "",
            r.destination or "",
        ]
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center

        data = files_map.get(r.source_file or "")
        if data:
            try:
                png_bytes = make_full_image_png_bytes(r.source_file or "", data)
                xl_img = XLImage(io.BytesIO(png_bytes))
                xl_img.width = int(xl_img.width * IMAGE_SCALE)
                xl_img.height = int(xl_img.height * IMAGE_SCALE)
                anchor = f"{get_column_letter(IMAGE_COL_IDX)}{row_idx}"
                ws.add_image(xl_img, anchor)
                ws.row_dimensions[row_idx].height = max(
                    20, xl_img.height * PX_TO_POINTS
                )
            except Exception:
                ws.cell(row=row_idx, column=IMAGE_COL_IDX, value=r.source_file or "")
        else:
            ws.cell(row=row_idx, column=IMAGE_COL_IDX, value=r.source_file or "")

    total_row = len(receipts) + 2
    ws.cell(row=total_row, column=5, value="합계").font = Font(bold=True)
    total_cell = ws.cell(
        row=total_row, column=6, value=sum(r.amount for r in receipts)
    )
    total_cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _image_bytes_to_pdf(data: bytes) -> bytes:
    img = Image.open(io.BytesIO(data))
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    out = io.BytesIO()
    img.save(out, format="PDF")
    return out.getvalue()


def bundle_receipts_pdf(files: list[tuple[str, bytes]]) -> bytes:
    """업로드된 영수증(이미지/PDF) 전체를 하나의 증빙 PDF로 병합."""
    writer = PdfWriter()
    for filename, data in files:
        ext = Path(filename).suffix.lower()
        if ext == ".pdf":
            reader = PdfReader(io.BytesIO(data))
            for page in reader.pages:
                writer.add_page(page)
        else:
            pdf_bytes = _image_bytes_to_pdf(data)
            reader = PdfReader(io.BytesIO(pdf_bytes))
            for page in reader.pages:
                writer.add_page(page)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()

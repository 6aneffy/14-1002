from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Literal, Optional

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

from .schema import Receipt
from .thumbnail import make_full_image_png_bytes

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "templates" / "여비정산신청서_템플릿.xlsx"

BOX_EMPTY = "☐"
BOX_FILLED = "■"

VehicleChoice = Literal["이용안함", "이용"]
SettlementChoice = Literal["숙박비실비", "숙박비정액"]


@dataclass
class FormMeta:
    doc_no: str = ""
    written_date: Optional[date] = None
    department: str = ""
    trip_period: str = ""
    trip_type: str = ""
    budget_item: str = ""
    remarks: str = ""
    purpose: str = ""

    vehicle: Optional[VehicleChoice] = None
    settlement: Optional[SettlementChoice] = None

    traveler_name: str = ""
    affiliation: str = ""
    position: str = ""
    travel_grade: str = ""
    individual_period: str = ""

    per_diem_cash: int = 0
    meal_cash: int = 0
    lodging_fixed_cash: int = 0
    lodging_actual_cash: int = 0
    prepaid_cash: int = 0

    per_diem_card: int = 0
    meal_card: int = 0
    lodging_fixed_card: int = 0
    lodging_actual_card: int = 0
    prepaid_card: int = 0

    lodging_rows: list[dict] = field(default_factory=list)


def format_trip_period_kr(d: date) -> str:
    """출장기간 표기: 2026년 4월 28일"""
    return f"{d.year}년 {d.month}월 {d.day}일"


def sheet_title_for_travel_date(d: date | None) -> str:
    """시트 탭 이름 (엑셀 31자·금지문자 제한)."""
    if d is None:
        base = "일자미상"
    else:
        base = f"{d.year}년{d.month}월{d.day}일"
    return sanitize_sheet_title(base)


def sanitize_sheet_title(title: str) -> str:
    cleaned = re.sub(r'[\[\]\:\*\?\/\\]', "", title)
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


def _render_choice(options: list[str], selected: Optional[str]) -> str:
    parts = []
    for opt in options:
        mark = BOX_FILLED if opt == selected else BOX_EMPTY
        parts.append(f"{mark} {opt}")
    return "    ".join(parts)


def _transport_label(r: Receipt) -> str:
    return r.transport.value


def _class_vehicle(r: Receipt) -> str:
    return r.travel_class or ""


def yeobi_sheet_total(meta: FormMeta, receipts: list[Receipt]) -> int:
    """
    시트별 여비(총괄) 합계. 템플릿과 동일하게 I열 소계(I20) + K열 선지급(K20).

    I18 = C18~G18(일비·식비·숙박정액·숙박실비·운임), I19는 카드 행, I20 = I18+I19.
    K20 = K18+K19(선지급 현금·카드).
    """
    fare = sum(r.amount for r in receipts)
    i18 = (
        meta.per_diem_cash
        + meta.meal_cash
        + meta.lodging_fixed_cash
        + meta.lodging_actual_cash
        + fare
    )
    i19 = (
        meta.per_diem_card
        + meta.meal_card
        + meta.lodging_fixed_card
        + meta.lodging_actual_card
    )
    i20 = i18 + i19
    k20 = meta.prepaid_cash + meta.prepaid_card
    return i20 + k20


def _ordered_receipt_files(
    receipts: list[Receipt], files_map: dict[str, bytes]
) -> list[tuple[str, bytes]]:
    """해당 일자 영수증 중, 업로드 원본이 있는 파일만 순서대로(파일명 중복 제거)."""
    out: list[tuple[str, bytes]] = []
    seen: set[str] = set()
    for r in receipts:
        fn = r.source_file or ""
        if not fn or fn in seen:
            continue
        data = files_map.get(fn)
        if not data:
            continue
        seen.add(fn)
        out.append((fn, data))
    return out


# 여비정산서: 행 높이(row_dimensions)를 바꾸지 않음 → 왼쪽 표(A~L) 줄 높이 유지.
# 이미지는 셀 위에 떠 있으므로 정산리스트(1/3)보다 크게 허용.
_SETTLEMENT_IMAGE_SCALE = 0.5
_SETTLEMENT_IMG_MAX_W = 380
_SETTLEMENT_IMG_MAX_H = 520
# 기본 행 높이(대략 px)로 다음 앵커 행 간격만 잡음 — 실제 행 높이는 수정하지 않음
_ROW_PX_EST = 18


def _fit_settlement_receipt_image(xl_img: XLImage) -> None:
    """여비정산서 전용 축소(정산리스트보다 큼) 후, 최대 픽셀만 한도(비율 유지)."""
    xl_img.width = int(xl_img.width * _SETTLEMENT_IMAGE_SCALE)
    xl_img.height = int(xl_img.height * _SETTLEMENT_IMAGE_SCALE)
    mw, mh = _SETTLEMENT_IMG_MAX_W, _SETTLEMENT_IMG_MAX_H
    if xl_img.width > mw:
        r = mw / xl_img.width
        xl_img.width = int(xl_img.width * r)
        xl_img.height = int(xl_img.height * r)
    if xl_img.height > mh:
        r = mh / xl_img.height
        xl_img.width = int(xl_img.width * r)
        xl_img.height = int(xl_img.height * r)


def _embed_receipt_images_column_n(
    ws,
    receipts: list[Receipt],
    files_map: dict[str, bytes],
    *,
    start_row: int = 4,
    col_letter: str = "N",
) -> None:
    """N4부터 아래로. 행 높이는 건드리지 않아 왼쪽 양식 표가 늘어나지 않음."""
    row = start_row
    for fn, data in _ordered_receipt_files(receipts, files_map):
        try:
            png_bytes = make_full_image_png_bytes(fn, data)
            xl_img = XLImage(io.BytesIO(png_bytes))
            _fit_settlement_receipt_image(xl_img)
            ws.add_image(xl_img, f"{col_letter}{row}")
            # 다음 영수증은 겹침 줄이기(행 자체 높이는 그대로)
            skip = max(2, int(float(xl_img.height) / _ROW_PX_EST) + 1)
            row += skip
        except Exception:
            continue


def _fill_settlement_sheet(
    ws,
    meta: FormMeta,
    receipts: list[Receipt],
    files_map: dict[str, bytes] | None = None,
) -> None:
    """단일 시트에 여비정산신청서 양식 채움."""
    if meta.written_date:
        ws["J3"] = meta.written_date.isoformat()
    if meta.doc_no:
        ws["C3"] = meta.doc_no

    ws["C10"] = meta.department
    ws["J10"] = meta.trip_period
    ws["C11"] = meta.trip_type
    ws["J11"] = meta.budget_item

    ws["C12"] = _render_choice(["이용안함", "이용"], meta.vehicle)
    ws["J12"] = _render_choice(["숙박비실비", "숙박비정액"], meta.settlement)

    ws["C13"] = meta.remarks
    ws["C14"] = meta.purpose
    ws["C14"].alignment = _LEFT

    ws["C18"] = meta.per_diem_cash
    ws["D18"] = meta.meal_cash
    ws["E18"] = meta.lodging_fixed_cash
    ws["F18"] = meta.lodging_actual_cash
    ws["K18"] = meta.prepaid_cash

    ws["C19"] = meta.per_diem_card
    ws["D19"] = meta.meal_card
    ws["E19"] = meta.lodging_fixed_card
    ws["F19"] = meta.lodging_actual_card
    ws["K19"] = meta.prepaid_card

    fare_total = sum(r.amount for r in receipts)
    ws["G18"] = fare_total

    # 수식 셀을 직접 계산값으로 덮어쓰기 → Protected View에서도 값이 표시됨
    i18 = (
        meta.per_diem_cash
        + meta.meal_cash
        + meta.lodging_fixed_cash
        + meta.lodging_actual_cash
        + fare_total
    )
    i19 = (
        meta.per_diem_card
        + meta.meal_card
        + meta.lodging_fixed_card
        + meta.lodging_actual_card
    )
    i20 = i18 + i19
    c20 = meta.per_diem_cash + meta.per_diem_card
    d20 = meta.meal_cash + meta.meal_card
    e20 = meta.lodging_fixed_cash + meta.lodging_fixed_card
    f20 = meta.lodging_actual_cash + meta.lodging_actual_card
    g20 = fare_total
    k20 = meta.prepaid_cash + meta.prepaid_card

    # 행18~20 소계 수식 셀
    ws["I18"] = i18
    ws["I19"] = i19
    ws["I20"] = i20
    ws["C20"] = c20
    ws["D20"] = d20
    ws["E20"] = e20
    ws["F20"] = f20
    ws["G20"] = g20
    ws["K20"] = k20

    # 행27 참조 수식 셀 (=E20, =F20, =G20, =I20, =K20)
    ws["C27"] = e20
    ws["D27"] = f20
    ws["E27"] = g20
    ws["G27"] = i20
    ws["I27"] = k20

    # 숙박 합계 (G34 = SUM(G31:G33))
    lodging_amount_total = sum(
        row.get("amount", 0) or 0 for row in meta.lodging_rows[:3]
    )
    ws["G34"] = lodging_amount_total

    # 영수증 운임 합계 (H43 = SUM(H38:H42))
    ws["H43"] = fare_total

    ws["A26"] = meta.traveler_name
    ws["B26"] = meta.affiliation
    ws["F26"] = meta.position
    ws["G26"] = meta.travel_grade
    ws["H26"] = meta.individual_period
    for coord in ("A26", "B26", "F26", "G26", "H26"):
        ws[coord].alignment = _CENTER

    for idx, row in enumerate(meta.lodging_rows[:3]):
        r = 31 + idx
        ws[f"A{r}"] = row.get("date", "")
        ws[f"B{r}"] = row.get("kind", "")
        ws[f"C{r}"] = row.get("name", "")
        ws[f"E{r}"] = row.get("industry", "")
        ws[f"F{r}"] = row.get("cancelled", "")
        ws[f"G{r}"] = row.get("amount", "")
        ws[f"H{r}"] = row.get("payment", "")
        ws[f"I{r}"] = row.get("approval_no", "")
        ws[f"K{r}"] = row.get("note", "")

    for idx, r in enumerate(receipts[:5]):
        row = 38 + idx
        ws[f"A{row}"] = r.travel_date.isoformat() if r.travel_date else ""
        ws[f"B{row}"] = _transport_label(r)
        ws[f"C{row}"] = r.origin or ""
        ws[f"D{row}"] = r.destination or ""
        ws[f"E{row}"] = _class_vehicle(r)
        ws[f"G{row}"] = ""
        ws[f"H{row}"] = r.amount
        ws[f"I{row}"] = ""
        ws[f"J{row}"] = "개인카드"
        ws[f"K{row}"] = ""
        ws[f"L{row}"] = meta.traveler_name

    _embed_receipt_images_column_n(ws, receipts, files_map or {})

    # L20은 템플릿에 수식/0이 섞일 수 있어, 합계 표시는 G27에 숫자로 기록 (엑셀에서 I20+K20과 일치)
    ws["G27"] = yeobi_sheet_total(meta, receipts)


def build_settlement_workbook(
    entries: list[tuple[str, FormMeta, list[Receipt]]],
    files_map: dict[str, bytes] | None = None,
) -> bytes:
    """
    운임일자(출장일)별 시트. 각 항목: (시트탭이름, 해당 일자용 FormMeta, 해당 일자 영수증만).

    템플릿 시트를 복제해 동일 양식을 채운다.
    """
    if not entries:
        raise ValueError("저장할 시트가 없습니다. 운임일자가 있는 영수증을 확인하세요.")

    wb = load_workbook(TEMPLATE_PATH)
    base = wb.worksheets[0]
    need = len(entries)
    while len(wb.worksheets) < need:
        wb.copy_worksheet(wb.worksheets[0])

    fm = files_map or {}
    for i, (title, meta, recs) in enumerate(entries):
        ws = wb.worksheets[i]
        ws.title = sanitize_sheet_title(title)
        _fill_settlement_sheet(ws, meta, recs, fm)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

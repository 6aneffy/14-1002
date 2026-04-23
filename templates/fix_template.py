"""템플릿에 개인별 지급여비 데이터 행(26행)을 추가하고
이동된 합계 수식을 보정한다. 1회성 수정 스크립트."""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

TEMPLATE = Path(__file__).parent / "여비정산신청서_템플릿.xlsx"

wb = load_workbook(TEMPLATE)
ws = wb.active

thin = Side(border_style="thin", color="000000")
BORDER = Border(top=thin, left=thin, right=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
FONT = Font(name="맑은 고딕", size=10)

# 새 26행 삽입 (라벨 25행과 기존 데이터 26행 사이)
ws.insert_rows(26)

# 새 26행: 성명/소속/직급/여비등급/출장기간의 데이터 셀 병합 설정
ws.merge_cells("B26:E26")
ws.merge_cells("H26:L26")
for coord in ("A26", "B26", "F26", "G26", "H26"):
    c = ws[coord]
    c.value = None
    c.alignment = CENTER
    c.font = FONT

# 26행 전체에 테두리 적용 (병합 안쪽 셀 포함)
for col in range(1, 13):
    cell = ws.cell(row=26, column=col)
    cell.border = BORDER

# 행 높이
ws.row_dimensions[26].height = 24

# 이동된 합계 수식 보정
# 기존: 숙박 합계 G33 = SUM(G30:G32) → 이동 후 G34 = SUM(G31:G33)
# 기존: 운임 합계 H42 = SUM(H37:H41) → 이동 후 H43 = SUM(H38:H42)
old_to_new = {
    "G34": "=SUM(G31:G33)",
    "H43": "=SUM(H38:H42)",
}
for coord, formula in old_to_new.items():
    print(f"  {coord} 기존값: {ws[coord].value!r} → {formula}")
    ws[coord] = formula

wb.save(TEMPLATE)
print(f"saved: {TEMPLATE}")
